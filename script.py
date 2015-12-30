import requests
from openpyxl import load_workbook, Workbook

def coordinates(address, api_key = None):
    url = 'https://maps.googleapis.com/maps/api/geocode/json'

    params = {
        'address' : address.encode('ascii', 'xmlcharrefreplace'),
        'sensor' : 'false'
    }

    if api_key:
        params['key'] = api_key

    response = requests.get(url, params=params)
    coord = response.json()

    if coord['status'] == 'OVER_QUERY_LIMIT':
        raise RuntimeError(coord['error_message'])

    if coord['status'] == 'OK':
        return {
            'lat': coord['results'][0]['geometry']['location']['lat'],
            'lng': coord['results'][0]['geometry']['location']['lng'],
        }
    else:
        return {
            'lat': '',
            'lng': '',
            }

def xlsx_coordinates(input_file, sheet_name, address_column, output_file, ignore_first_row = True, api_key = None):
    """
    Read the xlsx input_file, and add two new columns in the beginning of it with latitude and longitude
    Input file must be a xls* file
    """
    
    wb_read = load_workbook(input_file)
    sh = wb_read.get_sheet_by_name(sheet_name)
    if ignore_first_row:
        #Ignore first line of the file
        first = 1
    else:
        first = 0

    out = []
    for row in sh.rows[first:]:
        coord = coordinates(row[address_column].value, api_key)
        new_row = [coord['lat'],coord['lng']]
        for c in row:
            new_row.append(c.value)
        out.append(new_row)

    wb_write = Workbook(optimized_write=True)
    sheet = wb_write.create_sheet(sheet_name,0)

    for row in out:
        sheet.append(row)

    wb_write.save(output_file)

def create_kml(input_file,sheet_name,output_file,lat_column = 0,lng_column = 1):

    kml = simplekml.Kml()
    wb_read = load_workbook(input_file)
    sh = wb_read.get_sheet_by_name(sheet_name)

    properties = []
    for c in sh.rows[0]:
        properties.append(c.value.encode('utf-8'))

    for p,row in enumerate(sh.rows[1:]):
        for k,c in enumerate(row):
            if k==0:
                coord_tuple = (row[lng_column].value,row[lat_column].value)
                pnt = kml.newpoint(name = 'Point %s' % p, coords =[coord_tuple])
#            TODO: It's not working with unicode and utf-8
#            if k != lat_column and k != lng_column:
#                if type(row[k].value) == unicode or type(row[k].value) == str:
#                    pnt.extendeddata.newdata(properties[k],row[k].value.encode())

    kml.save(output_file)